from datetime import date

import pytest
from openpyxl import Workbook

from legacy_import.errors import WorkbookValidationError
from legacy_import.ids import stable_uuid
from legacy_import.values import duration_hours, parse_date
from legacy_import.wbs import nearest_ancestor_component
from legacy_import.workbook import LegacyWorkbook, SourceRow


def _save_workbook(tmp_path, rows):
    path = tmp_path / "legacy.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Source"
    worksheet.append(["ID", "Name"])
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    return path


def test_legacy_value_parsers_support_power_apps_formats() -> None:
    assert duration_hours("02:30") == 2.5
    assert parse_date("Wed Aug 27 2025 19:00:00 GMT-0500 (Peru Standard Time)") == date(
        2025,
        8,
        27,
    )
    assert parse_date("06/16") == date(2016, 6, 1)


def test_stable_uuid_is_repeatable_and_context_sensitive() -> None:
    first = stable_uuid("WBS_V2", "tbl_Component", 1, "assets")
    assert first == stable_uuid("WBS_V2", "tbl_Component", 1, "assets")
    assert first != stable_uuid("WBS_V2", "tbl_Equipment", 1, "assets")


def test_workbook_skips_identical_duplicate_source_rows(tmp_path) -> None:
    path = _save_workbook(tmp_path, [(1, "Equipo"), (1, "Equipo")])

    workbook = LegacyWorkbook(path)
    rows = workbook.rows("Source", key_column="ID")

    assert [row.key for row in rows] == ["1"]
    assert workbook.duplicate_counts["Source"] == 1


def test_workbook_rejects_conflicting_duplicate_source_rows(tmp_path) -> None:
    path = _save_workbook(tmp_path, [(1, "Equipo A"), (1, "Equipo B")])

    workbook = LegacyWorkbook(path)

    with pytest.raises(WorkbookValidationError, match="conflicting rows"):
        workbook.rows("Source", key_column="ID")


def _source_row(sheet: str, key: str, **values) -> SourceRow:
    return SourceRow(
        sheet=sheet,
        row_number=int(key),
        key=key,
        values=values,
        row_hash=key,
    )


def test_component_parent_uses_nearest_occupied_ancestor_slot() -> None:
    cabinet = _source_row(
        "tbl_SlotLocation",
        "137",
        FK_FatherSlotLocation=None,
    )
    pcsg_slot = _source_row(
        "tbl_SlotLocation",
        "148",
        FK_FatherSlotLocation=137,
    )
    cals_slot = _source_row(
        "tbl_SlotLocation",
        "158",
        FK_FatherSlotLocation=148,
    )
    pcsg = _source_row(
        "tbl_Component",
        "685",
        FK_CurrentLocation=9,
        FK_SlotLocation=148,
    )
    cals = _source_row(
        "tbl_Component",
        "693",
        FK_CurrentLocation=9,
        FK_SlotLocation=158,
    )

    parent = nearest_ancestor_component(
        cals,
        slots_by_key={
            cabinet.key: cabinet,
            pcsg_slot.key: pcsg_slot,
            cals_slot.key: cals_slot,
        },
        components_by_location_and_slot={("9", "148"): pcsg},
    )

    assert parent == pcsg


def test_component_without_occupied_ancestor_slot_uses_equipment_fallback() -> None:
    cabinet = _source_row(
        "tbl_SlotLocation",
        "137",
        FK_FatherSlotLocation=None,
    )
    pcsg_slot = _source_row(
        "tbl_SlotLocation",
        "148",
        FK_FatherSlotLocation=137,
    )
    pcsg = _source_row(
        "tbl_Component",
        "685",
        FK_CurrentLocation=9,
        FK_SlotLocation=148,
    )

    parent = nearest_ancestor_component(
        pcsg,
        slots_by_key={
            cabinet.key: cabinet,
            pcsg_slot.key: pcsg_slot,
        },
        components_by_location_and_slot={},
    )

    assert parent is None
