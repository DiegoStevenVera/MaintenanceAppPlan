import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from legacy_import.errors import WorkbookValidationError
from legacy_import.values import identifier, json_value


@dataclass(frozen=True)
class SourceRow:
    sheet: str
    row_number: int
    key: str
    values: dict[str, Any]
    row_hash: str


class LegacyWorkbook:
    def __init__(self, path: Path) -> None:
        if not path.is_file():
            raise WorkbookValidationError(f"Workbook not found: {path}")
        self.path = path.resolve()
        self.file_checksum = self._checksum(self.path)
        self._workbook = load_workbook(self.path, read_only=True, data_only=True)
        self.duplicate_counts: dict[str, int] = {}

    @property
    def sheet_names(self) -> tuple[str, ...]:
        return tuple(self._workbook.sheetnames)

    def rows(
        self,
        sheet_name: str,
        *,
        key_column: str,
        required_columns: set[str] | None = None,
    ) -> list[SourceRow]:
        if sheet_name not in self._workbook.sheetnames:
            raise WorkbookValidationError(f"Missing worksheet: {sheet_name}")

        worksheet = self._workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration as error:
            raise WorkbookValidationError(f"Worksheet is empty: {sheet_name}") from error

        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        if key_column not in headers:
            raise WorkbookValidationError(
                f"{sheet_name} does not contain primary key column {key_column!r}"
            )
        missing = (required_columns or set()) - set(headers)
        if missing:
            raise WorkbookValidationError(
                f"{sheet_name} is missing columns: {', '.join(sorted(missing))}"
            )

        rows_by_key: dict[str, SourceRow] = {}
        duplicates = 0
        for row_number, raw_row in enumerate(iterator, start=2):
            values = {
                header: raw_row[index] if index < len(raw_row) else None
                for index, header in enumerate(headers)
                if header
            }
            if not any(value is not None and str(value).strip() for value in values.values()):
                continue
            try:
                source_key = identifier(values.get(key_column))
            except ValueError as error:
                raise WorkbookValidationError(
                    f"{sheet_name} row {row_number}: {error}"
                ) from error
            row_hash = self.hash_values(values)
            source_row = SourceRow(
                sheet=sheet_name,
                row_number=row_number,
                key=source_key,
                values=values,
                row_hash=row_hash,
            )
            previous = rows_by_key.get(source_key)
            if previous is None:
                rows_by_key[source_key] = source_row
            elif previous.row_hash == row_hash:
                duplicates += 1
            else:
                raise WorkbookValidationError(
                    f"{sheet_name} has conflicting rows for key {source_key!r}: "
                    f"rows {previous.row_number} and {row_number}"
                )

        self.duplicate_counts[sheet_name] = duplicates
        return list(rows_by_key.values())

    @staticmethod
    def hash_values(values: dict[str, Any]) -> str:
        normalized = {
            key: json_value(value)
            for key, value in sorted(values.items())
        }
        payload = json.dumps(
            normalized,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    @staticmethod
    def _checksum(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
